"""Tests for Excel extraction module."""

import sys
import os
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from extractors.excel_extractor import ExcelExtractor


class TestExcelExtractor:

    def setup_method(self):
        self.extractor = ExcelExtractor()

    def test_extract_rows_nonexistent_file(self):
        """Should return empty list for a file that doesn't exist."""
        rows = self.extractor.extract_rows("/nonexistent/path/file.xlsx")
        assert rows == []

    def test_extract_rows_corrupt_file(self):
        """Should handle a corrupt/non-Excel file gracefully."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"not a real excel file")
            tmp_path = f.name

        try:
            rows = self.extractor.extract_rows(tmp_path)
            assert isinstance(rows, list)
            assert rows == []
        finally:
            os.unlink(tmp_path)

    def test_extract_rows_valid_excel(self):
        """Should parse a valid Excel file into list of dicts."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Name", "Value", "Status"])
            ws.append(["Facility A", 1000, "Active"])
            ws.append(["Facility B", 2000, "Inactive"])
            wb.save(tmp_path)
            wb.close()

            rows = self.extractor.extract_rows(tmp_path)
            assert len(rows) == 2
            assert rows[0]["Name"] == "Facility A"
            assert rows[0]["Value"] == 1000
            assert rows[1]["Status"] == "Inactive"
        finally:
            os.unlink(tmp_path)

    def test_extract_rows_header_only(self):
        """Should return empty list when Excel has only a header row."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Name", "Value"])
            wb.save(tmp_path)
            wb.close()

            rows = self.extractor.extract_rows(tmp_path)
            assert rows == []
        finally:
            os.unlink(tmp_path)

    def test_extract_rows_none_headers(self):
        """Should generate col_N names for None header cells."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Name", None, "Status"])
            ws.append(["Test", 42, "OK"])
            wb.save(tmp_path)
            wb.close()

            rows = self.extractor.extract_rows(tmp_path)
            assert len(rows) == 1
            assert "col_1" in rows[0]
            assert rows[0]["col_1"] == 42
        finally:
            os.unlink(tmp_path)
